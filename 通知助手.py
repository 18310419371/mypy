#开发日志2020.4.7需求检查报平安完成情况并发到微信群
#开发日志2020.4.8进行python源代码打包
#开发日志2020.4.9改进微信输出方式多条输出成一条
#开发日志2020.4.9现需增加功能，待测数据样例转换为UTF-8的测试样例
#开发日志2020.4.9新加功能文件转换，输入网站上下载的文件名进行数据转换 修复bug提高使用效率
#开发日志2020.4.9现需增加功能，配置文件读取
#开发日志2020.4.10新加功能群名文件读取，不使用微信群时进行文件生成
#开发日志2020.4.11成功在未安装python环境下运行打包程序
#开发日志2020.4.12新增可编辑说话
import csv
from openpyxl import load_workbook
from wxpy import *
from time import sleep
############################################################################################

#找到需要xlsx文件的位置
name=input("请输入要转换的文件名:")
writedata = open("data/test.csv", "w", encoding="UTF-8-sig")#新建转换后文件对象
workbook = load_workbook(name+".xlsx")#打开要转换文件
#获取当前活跃的sheet,默认是第一个sheet
booksheet = workbook.active
#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns
i = 0
x=0
for row in rows:
    i = i + 1
    line = [col.value for col in row]
    for j in range(1,20):
        if(booksheet.cell(row=i, column=j).value=="姓名"):
            x=j
            break
    break
workbook.close()

workbook = load_workbook(name+".xlsx")
#获取当前活跃的sheet,默认是第一个sheet
booksheet = workbook.active
#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns
writedata_str = ""
# 迭代所有的行
i=0
for row in rows:
    i = i + 1
    line = [col.value for col in row]
    writedata_str+=booksheet.cell(row=i, column=x).value+",\n"
writedata.write(writedata_str)#文件写入数据
writedata.close()
workbook.close()
print("转换成功！\n")
print("即将进行审查！\n")
sleep(2)
############################################################################################

############################################################################################
name_all = []
name_test = []
name_loss = []
flag = 0
all = open("data/人名单.csv", "r", encoding="UTF-8-sig")
test = open("data/test.csv", "r", encoding="UTF-8-sig")
allname = csv.reader(all)
testname = csv.reader(test)
for row in testname:
    name_test.append(row[0])#遍历测试对象写入测试人名数组
for row in allname:
    name_all.append(row[0])
all.close()
test.close()
for i in name_all:    #总人数要改
    for j in name_test:
        if(i==j):
            flag=1
    if(flag==0):
        name_loss.append(i)
    else:
        flag=0
str = ""
i=0
for lossname in name_loss:
    i+=1
    str+=lossname+"\n"
print("审查完毕！")
print("共有",i,"个未提交\n")
weixin=input("请选择是否发送到微信群Y/N否则生成名单文件:")
if(weixin=="Y" or weixin=="y"):
    print("\n请扫码进行登录！")
    sleep(2)
    bot = Bot()
    myclass=open("data/微信群名.txt",encoding="UTF-8-sig")
    mygroup = bot.groups().search(myclass.read())[0]  #微信群名  要改
    # mygroup = bot.file_helper
    myspeak = open("data/要说的话.txt", encoding="UTF-8-sig")
    strspeak = myspeak.read()
    myspeak.close()
    mygroup.send(str+strspeak)
    myclass.close()
    print("运行完毕，即将退出")
else:
    a = open("name.txt", "w", encoding="UTF-8-sig")
    a.write(str)
    a.close()
    print("名单以生成（name.txt）\n运行完毕，即将退出")
sleep(3)
# print(str+"请以上同学报平安")   #调试用
############################################################################################
